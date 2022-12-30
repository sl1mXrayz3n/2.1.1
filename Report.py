import openpyxl
from openpyxl.styles import Side, Border, Font


class Report:
    def __init__(self, salary_levels_by_years: dict,
                 number_vacancies_by_years: dict,
                 salaries_years_chosen_profession: dict,
                 number_vacancies_years_chosen_profession: dict,
                 salary_by_city: dict,
                 salary_levels_by_city: dict,
                 profession: str):

        self.__salary_levels_by_years = salary_levels_by_years
        self.__number_vacancies_by_years = number_vacancies_by_years
        self.__salaries_years_chosen_profession = salaries_years_chosen_profession
        self.__number_vacancies_years_chosen_profession = number_vacancies_years_chosen_profession
        self.__salary_by_city = salary_by_city
        self.__salary_levels_by_city = salary_levels_by_city
        self.__profession = profession

    def __generate_statistic_years(self, book, sheet_name, index):
        book.create_sheet(sheet_name)
        title = ["Год",
                 "Средняя зарплата",
                 f"Средняя зарплата - {self.__profession}",
                 "Количество вакансий",
                 f"Количество вакансий - {self.__profession}"]
        book.worksheets[index].append(title)
        for year in self.__salary_levels_by_years.keys():
            book.worksheets[index].append([year,
                                           self.__salary_levels_by_years[year],
                                           self.__salaries_years_chosen_profession[year],
                                           self.__number_vacancies_by_years[year],
                                           self.__number_vacancies_years_chosen_profession[year]])

        for i in range(len(title)):
            book.worksheets[index].cell(1, i + 1).font = Font(bold=True)

        side = Side(border_style='thin', color="FF000000")
        border = Border(
            left=side,
            right=side,
            top=side,
            bottom=side
        )
        for i in range(len(self.__salary_levels_by_years.keys()) + 1):
            for j in range(len(title)):
                book.worksheets[index].cell(i + 1, j + 1).border = border

        dimensions = {}
        for row in book.worksheets[index].rows:
            for cell in row:
                if cell.value:
                    dimensions[cell.column_letter] = max((dimensions.get(cell.column_letter, 0),
                                                          len(str(cell.value)) + 2))
        for col, value in dimensions.items():
            book.worksheets[index].column_dimensions[col].width = value

    def __generate_statistic_cities(self, book, sheet_name, index):
        book.create_sheet(sheet_name)
        side = Side(border_style='thin', color="FF000000")
        border = Border(
            left=side,
            right=side,
            top=side,
            bottom=side
        )

        book.create_sheet("Статистика по городам")
        title = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]

        book.worksheets[index].append(title)

        cities1 = list(self.__salary_by_city.keys())
        cities2 = list(self.__salary_levels_by_city.keys())
        for i in range(len(cities1)):
            book.worksheets[index].append([cities1[i],
                                           self.__salary_by_city[cities1[i]],
                                           "",
                                           cities2[i],
                                           self.__salary_levels_by_city[cities2[i]]])

        for i in range(len(title)):
            book.worksheets[index].cell(1, i + 1).font = Font(bold=True)

        for i in range(len(cities1) + 1):
            for j in range(len(title)):
                book.worksheets[index].cell(i + 1, j + 1).border = border

        for i in range(2, len(cities2) + 2):
            book.worksheets[index].cell(i, 5).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[10]

        dimensions = {}
        for row in book.worksheets[1].rows:
            for cell in row:
                if cell.value:
                    dimensions[cell.column_letter] = max(
                        (dimensions.get(cell.column_letter, 0), len(str(cell.value)) + 2))
        for col, value in dimensions.items():
            book.worksheets[index].column_dimensions[col].width = value

    def generate_excel(self):
        book = openpyxl.Workbook()
        book.remove(book["Sheet"])
        self.__generate_statistic_years(book, "Статистика по годам", 0)
        self.__generate_statistic_cities(book, "Статистика по городам", 1)
        book.save("report.xlsx")